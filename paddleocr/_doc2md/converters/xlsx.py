# Copyright (c) 2025 PaddlePaddle Authors. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
from io import BytesIO
from pathlib import Path
from typing import Optional

from ..base import BaseConverter, ConvertResult
from ..registry import default_registry


def _get_sheet_width_emu(ws, openpyxl_mod) -> int:
    """Return the total column width of the worksheet in EMU. 1 char width ~ 7px, 1px = 9525 EMU."""
    CHAR_TO_EMU = 7 * 9525
    total = 0
    for col_idx in range(1, (ws.max_column or 1) + 1):
        col_letter = openpyxl_mod.utils.get_column_letter(col_idx)
        col_dim = ws.column_dimensions.get(col_letter)
        if col_dim and col_dim.width is not None:
            total += col_dim.width * CHAR_TO_EMU
        else:
            total += 8.43 * CHAR_TO_EMU  # default column width
    return int(total)


def _get_image_cx(anchor) -> Optional[int]:
    """Return image display width in EMU for OneCellAnchor; return None for TwoCellAnchor."""
    try:
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor

        if isinstance(anchor, OneCellAnchor):
            return anchor.ext.cx
    except (AttributeError, ImportError):
        pass
    return None


def _find_data_bounds(ws, image_map, max_rows=None):
    """Return (min_row, max_row, min_col, max_col) 1-based, or None if the sheet is empty."""
    min_r = min_c = float("inf")
    max_r = max_c = 0

    # Non-empty cells
    for cell in ws._cells.values():
        if cell.value is not None:
            r, c = cell.row, cell.column
            if max_rows is not None and r > max_rows:
                continue
            min_r = min(min_r, r)
            max_r = max(max_r, r)
            min_c = min(min_c, c)
            max_c = max(max_c, c)

    # Merged cell ranges
    for mr in ws.merged_cells.ranges:
        r1, r2 = mr.min_row, mr.max_row
        if max_rows is not None:
            r2 = min(r2, max_rows)
        if r1 > r2:
            continue
        min_r = min(min_r, r1)
        max_r = max(max_r, r2)
        min_c = min(min_c, mr.min_col)
        max_c = max(max_c, mr.max_col)

    # Image anchors (0-based -> 1-based)
    for img_r0, img_c0 in image_map:
        r, c = img_r0 + 1, img_c0 + 1
        if max_rows is not None and r > max_rows:
            continue
        min_r = min(min_r, r)
        max_r = max(max_r, r)
        min_c = min(min_c, c)
        max_c = max(max_c, c)

    if max_r == 0:
        return None
    return (int(min_r), int(max_r), int(min_c), int(max_c))


@default_registry.register
class XlsxConverter(BaseConverter):
    supported_extensions = ["xlsx"]
    supported_mimetypes = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ]

    def convert_file(self, file_path: Path, **kwargs) -> ConvertResult:
        try:
            import openpyxl
            from openpyxl.drawing.spreadsheet_drawing import (
                OneCellAnchor,
                TwoCellAnchor,
            )
        except ImportError:
            raise RuntimeError(
                "XLSX conversion requires openpyxl: pip install paddleocr[doc2md]"
            )

        sheet_name: Optional[str] = kwargs.get("sheet_name", None)
        max_rows: Optional[int] = kwargs.get("max_rows", None)

        # read_only=False is required to access merged_cells
        wb = openpyxl.load_workbook(str(file_path), read_only=False, data_only=True)
        sheets_md = []
        images: dict = {}
        image_counter = 0

        target_sheets = [sheet_name] if sheet_name else wb.sheetnames

        for sname in target_sheets:
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]

            # Total sheet width in EMU, used for image percentage calculation
            sheet_width_emu = _get_sheet_width_emu(ws, openpyxl)

            # Floating image map: (0-based row, 0-based col) -> [Image, ...]
            image_map: dict = {}
            for img in getattr(ws, "_images", []):
                anchor = img.anchor
                if isinstance(anchor, (OneCellAnchor, TwoCellAnchor)):
                    r, c = anchor._from.row, anchor._from.col
                    image_map.setdefault((r, c), []).append(img)
                elif isinstance(anchor, str):
                    try:
                        from openpyxl.utils import coordinate_to_tuple

                        r, c = coordinate_to_tuple(anchor)
                        image_map.setdefault((r - 1, c - 1), []).append(img)
                    except Exception:
                        pass

            # Merged cell map: (row, col) -> MergedCellRange
            merge_map = {}
            for mr in ws.merged_cells.ranges:
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        merge_map[(r, c)] = mr

            # Trim surrounding empty rows/columns
            bounds = _find_data_bounds(ws, image_map, max_rows)
            if bounds is None:
                continue
            data_min_row, data_max_row, data_min_col, data_max_col = bounds

            html_parts = ["<table>"]
            for row_idx in range(data_min_row, data_max_row + 1):
                html_parts.append("<tr>")
                for col_idx in range(data_min_col, data_max_col + 1):
                    cell = ws.cell(row_idx, col_idx)
                    mr = merge_map.get((row_idx, col_idx))
                    # Skip non-origin cells in a merged range
                    if mr and (row_idx, col_idx) != (mr.min_row, mr.min_col):
                        continue
                    tag = "th" if row_idx == data_min_row else "td"
                    attrs = ""
                    if mr:
                        cs = (
                            min(mr.max_col, data_max_col)
                            - max(mr.min_col, data_min_col)
                            + 1
                        )
                        rs = (
                            min(mr.max_row, data_max_row)
                            - max(mr.min_row, data_min_row)
                            + 1
                        )
                        if cs > 1:
                            attrs += f' colspan="{cs}"'
                        if rs > 1:
                            attrs += f' rowspan="{rs}"'

                    # Cell text
                    value = cell.value
                    text = str(value) if value is not None else ""

                    # Floating images
                    cell_images = image_map.get((row_idx - 1, col_idx - 1), [])
                    img_html = ""
                    for img_obj in cell_images:
                        image_counter += 1
                        ext = (img_obj.format or "png").lower()
                        filename = f"image{image_counter}.{ext}"
                        rel_path = f"images/{filename}"
                        try:
                            ref = img_obj.ref
                            if isinstance(ref, BytesIO):
                                ref.seek(0)
                                images[rel_path] = ref.read()
                            else:
                                images[rel_path] = img_obj._data()
                            cx_emu = _get_image_cx(img_obj.anchor)
                            if cx_emu and sheet_width_emu:
                                pct = min(round(cx_emu / sheet_width_emu * 100), 100)
                                img_html += (
                                    f'<img src="images/{filename}" width="{pct}%">'
                                )
                            else:
                                img_html += f'<img src="images/{filename}">'
                        except Exception:
                            pass

                    cell_content = img_html + text if img_html else text
                    html_parts.append(f"<{tag}{attrs}>{cell_content}</{tag}>")
                html_parts.append("</tr>")
            html_parts.append("</table>")

            table_html = "\n".join(html_parts)
            sheets_md.append(f"## {sname}\n\n{table_html}")

        wb.close()

        md_text = "\n\n".join(sheets_md)

        return ConvertResult(
            markdown=md_text,
            images=images,
            metadata={
                "format": "XLSX",
                "sheet_count": len(wb.sheetnames),
                "sheet_names": list(wb.sheetnames),
            },
        )
